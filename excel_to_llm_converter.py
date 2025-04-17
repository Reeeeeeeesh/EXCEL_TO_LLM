import pandas as pd
import openpyxl
from pathlib import Path
import json
import re
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple, Any
from datetime import datetime
from combine_markdown import combine_markdown_files
from llm_analyzer import LLMAnalyzer  # Import LLMAnalyzer
import os

# Define the root directory
ROOT_DIR = r"C:\Users\sueho\Documents\EXCEL_TO_LLM"

# Define input and output paths
INPUT_DIR = Path(ROOT_DIR) / "INPUT"
OUTPUT_DIR = Path(ROOT_DIR) / "OUTPUT"

class ExcelToLLMConverter:
    def __init__(self, input_path: str, output_dir: str, api_key: str):
        self.input_path = Path(input_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.llm_analyzer = LLMAnalyzer(api_key)  # Initialize LLMAnalyzer

    def infer_cell_type(self, cell: openpyxl.cell.Cell) -> str:
        """Infer the type of data in a cell."""
        if cell.value is None:
            return "empty"
        
        if cell.data_type == 'n':  # Numeric
            if isinstance(cell.value, datetime):
                return "date"
            if str(cell.number_format).endswith('%'):
                return "percentage"
            if any(currency_symbol in str(cell.number_format) for currency_symbol in ['$', '£', '€', '¥']):
                return "currency"
            return "numeric"
        elif cell.data_type == 'f':
            return "formula"
        elif cell.data_type == 'b':
            return "boolean"
        elif cell.data_type == 's':
            return "text"
        
        return str(cell.data_type)

    def identify_tables(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Identify tables within the worksheet."""
        tables = []
        current_table = None
        
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            # Look for potential header rows (cells with different styling)
            header_candidates = [cell for cell in row if cell.font.bold or (cell.fill and cell.fill.start_color)]
            
            if header_candidates and len(header_candidates) > 1:
                # Found potential new table header
                if current_table:
                    current_table["row_count"] = row_idx - current_table["start_row"]
                    tables.append(current_table)
                
                # Start new table
                start_col = min(cell.column for cell in header_candidates)
                end_col = max(cell.column for cell in header_candidates)
                headers = []
                types = []
                
                for cell in row[start_col-1:end_col]:
                    headers.append(str(cell.value) if cell.value else f"Column_{get_column_letter(cell.column)}")
                    types.append(self.infer_cell_type(cell))
                
                current_table = {
                    "name": headers[0] if headers else f"Table_{len(tables)+1}",
                    "range": f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}",
                    "headers": headers,
                    "types": types,
                    "start_row": row_idx,
                    "start_col": start_col,
                    "end_col": end_col
                }
        
        # Close last table if exists
        if current_table:
            current_table["row_count"] = worksheet.max_row - current_table["start_row"] + 1
            tables.append(current_table)
        
        return tables

    def extract_named_ranges(self, workbook: openpyxl.Workbook) -> List[Dict[str, str]]:
        """Extract named ranges from the workbook."""
        named_ranges = []
        try:
            # Handle newer versions of openpyxl
            if hasattr(workbook, 'defined_names'):
                for name, defn in workbook.defined_names.items():
                    try:
                        destinations = defn.destinations
                        for worksheet, coordinate in destinations:
                            named_ranges.append({
                                "name": name,
                                "range": f"{worksheet}!{coordinate}"
                            })
                    except AttributeError:
                        # If destinations not available, try to get the value directly
                        named_ranges.append({
                            "name": name,
                            "range": str(defn.value)
                        })
        except Exception as e:
            print(f"Warning: Could not extract named ranges: {str(e)}")
        
        return named_ranges

    def identify_key_sections(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Identify key sections like P&L, Balance Sheet, etc."""
        key_sections = []
        section_keywords = {
            "P&L Statement": ["profit", "loss", "p&l", "income statement"],
            "Balance Sheet": ["balance sheet", "assets", "liabilities"],
            "Cash Flow": ["cash flow", "operating activities", "financing activities"]
        }
        
        current_section = None
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = cell.value.lower()
                    for section_name, keywords in section_keywords.items():
                        if any(keyword in cell_value for keyword in keywords):
                            if current_section:
                                current_section["end_row"] = row_idx - 1
                                key_sections.append(current_section)
                            
                            current_section = {
                                "name": section_name,
                                "start_row": row_idx,
                                "range": f"{get_column_letter(cell.column)}{row_idx}"
                            }
                            break
        
        if current_section:
            current_section["end_row"] = worksheet.max_row
            key_sections.append(current_section)
        
        return key_sections

    def categorize_formula(self, formula: str) -> str:
        """Categorize formula type."""
        formula = formula.upper()
        if '!' in formula:
            return "external_reference"
        elif any(agg in formula for agg in ["SUM(", "AVERAGE(", "COUNT(", "MAX(", "MIN("]):
            return "aggregation"
        elif "IF(" in formula:
            return "conditional_logic"
        return "other"

    def extract_formula_metadata(self, cell: openpyxl.cell.Cell) -> Dict[str, Any]:
        """Extract enhanced formula metadata."""
        metadata = {
            "address": f"{get_column_letter(cell.column)}{cell.row}",
            "value": cell.value,
            "data_type": self.infer_cell_type(cell),
            "formula": None,
            "category": None,
            "dependencies": []
        }

        if cell.data_type == 'f':
            formula = str(cell.value)
            if formula.startswith('='):
                metadata["formula"] = formula
                metadata["category"] = self.categorize_formula(formula)
                metadata["dependencies"] = re.findall(r'[A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?', formula)

        return metadata

    def process_worksheet(self, worksheet: Worksheet, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Process a worksheet and extract enhanced structure and metadata."""
        sheet_data = {
            "name": worksheet.title,
            "dimensions": f"{worksheet.dimensions}",
            "tables": self.identify_tables(worksheet),
            "named_ranges": self.extract_named_ranges(workbook),
            "key_sections": self.identify_key_sections(worksheet),
            "cells": {},
            "formulas": {
                "external_references": [],
                "aggregations": [],
                "conditional_logic": [],
                "other": []
            },
            "data_relationships": []
        }

        # Process each cell
        for row in worksheet.iter_rows():
            for cell in row:
                try:
                    if cell.value is not None:
                        metadata = self.extract_formula_metadata(cell)
                        cell_address = metadata["address"]

                        # Store basic cell data
                        sheet_data["cells"][cell_address] = {
                            "value": str(cell.value),
                            "type": metadata["data_type"]
                        }

                        # Store formula information if present
                        if metadata["formula"]:
                            formula_data = {
                                "address": cell_address,
                                "formula": metadata["formula"],
                                "dependencies": metadata["dependencies"]
                            }
                            sheet_data["formulas"][metadata["category"]].append(formula_data)

                            # Record data relationships
                            for dep in metadata["dependencies"]:
                                sheet_data["data_relationships"].append({
                                    "source": dep,
                                    "target": cell_address,
                                    "type": "formula_dependency"
                                })
                except Exception as e:
                    print(f"Error processing cell {get_column_letter(cell.column)}{cell.row}: {str(e)}")
                    continue

        return sheet_data

    def sanitize_filename(self, filename: str) -> str:
        """Sanitize filename by removing or replacing invalid characters."""
        # Characters that are not allowed in Windows filenames
        invalid_chars = '<>:"/\\|?*'
        # Replace invalid characters with underscore
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename.strip()

    def convert_to_markdown(self, sheet_data: Dict[str, Any], output_file: Path) -> None:
        """Convert sheet data to a markdown file optimized for LLM ingestion."""
        with output_file.open('w', encoding='utf-8') as f:
            # Write header
            f.write(f"# Sheet: {sheet_data['name']}\n\n")
            f.write(f"Dimensions: {sheet_data['dimensions']}\n\n")

            # Write tables section
            if sheet_data["tables"]:
                f.write("## Tables\n\n")
                for table in sheet_data["tables"]:
                    f.write(f"### {table['name']}\n")
                    f.write(f"- Range: {table['range']}\n")
                    f.write(f"- Headers: {', '.join(table['headers'])}\n")
                    f.write(f"- Types: {', '.join(table['types'])}\n")
                    f.write("\n")

            # Write named ranges section
            if sheet_data["named_ranges"]:
                f.write("## Named Ranges\n\n")
                for named_range in sheet_data["named_ranges"]:
                    f.write(f"- {named_range['name']}: {named_range['range']}\n")

            # Write key sections section
            if sheet_data["key_sections"]:
                f.write("## Key Sections\n\n")
                for section in sheet_data["key_sections"]:
                    f.write(f"- {section['name']}: {section['range']}\n")

            # Write cell values section
            f.write("\n## Cell Values\n\n")
            f.write("| Cell | Value | Type |\n")
            f.write("|------|--------|------|\n")
            for addr, cell_data in sheet_data["cells"].items():
                # Escape pipe characters in cell values
                safe_value = str(cell_data['value']).replace('|', '\\|')
                f.write(f"| {addr} | {safe_value} | {cell_data['type']} |\n")

            # Write formulas section
            if any(sheet_data["formulas"].values()):
                f.write("\n## Formulas\n\n")
                for category, formulas in sheet_data["formulas"].items():
                    if formulas:
                        f.write(f"### {category.capitalize()}\n")
                        for formula in formulas:
                            f.write(f"- {formula['address']}: `{formula['formula']}`\n")
                            if formula['dependencies']:
                                f.write(f"- Dependencies: {', '.join(formula['dependencies'])}\n")
                            f.write("\n")

            # Write data relationships section
            if sheet_data["data_relationships"]:
                f.write("\n## Data Relationships\n\n")
                for rel in sheet_data["data_relationships"]:
                    f.write(f"- {rel['source']} → {rel['target']} ({rel['type']})\n")

    def generate_workbook_summary(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Generate a high-level summary of the workbook."""
        summary = {
            "sheet_count": len(workbook.worksheets),
            "sheets": [],
            "formula_patterns": {},
            "most_formulas": {"sheet": None, "count": 0}
        }

        for worksheet in workbook.worksheets:
            sheet_summary = {
                "name": worksheet.title,
                "formula_count": 0,
                "table_count": len(self.identify_tables(worksheet)),
                "key_sections": [section["name"] for section in self.identify_key_sections(worksheet)]
            }

            # Count formulas and patterns
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        sheet_summary["formula_count"] += 1
                        formula = str(cell.value).upper()
                        for pattern in ["SUM(", "IF(", "VLOOKUP(", "INDEX(", "MATCH("]:
                            if pattern in formula:
                                summary["formula_patterns"][pattern] = summary["formula_patterns"].get(pattern, 0) + 1

            summary["sheets"].append(sheet_summary)
            if sheet_summary["formula_count"] > summary["most_formulas"]["count"]:
                summary["most_formulas"] = {
                    "sheet": worksheet.title,
                    "count": sheet_summary["formula_count"]
                }

        # Sort formula patterns by frequency
        summary["formula_patterns"] = dict(
            sorted(summary["formula_patterns"].items(), key=lambda x: x[1], reverse=True)
        )

        return summary

    def process_workbook(self, excel_file: Path) -> None:
        """Process an entire workbook and generate output files."""
        try:
            print(f"Processing {excel_file}...")
            workbook = openpyxl.load_workbook(excel_file, data_only=False)

            # Generate workbook summary
            workbook_summary = self.generate_workbook_summary(workbook)

            # Create output directory for this workbook
            workbook_dir = self.output_dir / excel_file.stem
            workbook_dir.mkdir(exist_ok=True)

            # Save workbook summary
            with (workbook_dir / "workbook_summary.md").open('w', encoding='utf-8') as f:
                f.write("# Workbook Summary\n\n")
                f.write(f"Total Sheets: {workbook_summary['sheet_count']}\n\n")
                
                f.write("## Sheet Details\n")
                for sheet in workbook_summary["sheets"]:
                    f.write(f"\n### {sheet['name']}\n")
                    f.write(f"- Formula Count: {sheet['formula_count']}\n")
                    f.write(f"- Table Count: {sheet['table_count']}\n")
                    if sheet['key_sections']:
                        f.write(f"- Key Sections: {', '.join(sheet['key_sections'])}\n")
                
                f.write("\n## Formula Analysis\n")
                f.write(f"Sheet with Most Formulas: {workbook_summary['most_formulas']['sheet']} ")
                f.write(f"({workbook_summary['most_formulas']['count']} formulas)\n\n")
                
                if workbook_summary["formula_patterns"]:
                    f.write("Common Formula Patterns:\n")
                    for pattern, count in workbook_summary["formula_patterns"].items():
                        f.write(f"- {pattern}: {count} occurrences\n")

            # Process each worksheet
            for worksheet in workbook.worksheets:
                print(f"Processing worksheet: {worksheet.title}")
                sheet_data = self.process_worksheet(worksheet, workbook)

                # Sanitize the worksheet title for filename
                safe_title = self.sanitize_filename(worksheet.title)
                if not safe_title:  # If title becomes empty after sanitization
                    safe_title = "Sheet"

                # Save as markdown
                md_file = workbook_dir / f"{safe_title}.md"
                self.convert_to_markdown(sheet_data, md_file)
                print(f"Created markdown file: {md_file}")

                # Save raw data as JSON for potential other uses
                json_file = workbook_dir / f"{safe_title}.json"
                with json_file.open('w', encoding='utf-8') as f:
                    json.dump(sheet_data, f, indent=2)
                print(f"Created JSON file: {json_file}")

        except Exception as e:
            print(f"Error processing {excel_file}: {str(e)}")

    def convert_all(self):
        """Convert all Excel files in the input path."""
        if self.input_path.is_file():
            self.process_workbook(self.input_path)
        else:
            for excel_file in self.input_path.glob("*.xlsx"):
                self.process_workbook(excel_file)
        
        # After processing all Excel files, combine the markdown files for each workbook directory
        for workbook_dir in self.output_dir.iterdir():
            if workbook_dir.is_dir():  # Process each workbook directory
                print(f"\nCombining markdown files for {workbook_dir.name}...")
                combined_file = combine_markdown_files(str(workbook_dir))
                
                # Read the combined markdown content
                if combined_file and os.path.exists(combined_file):
                    print(f"Successfully created combined markdown file: {combined_file}")
                    try:
                        with open(combined_file, 'r', encoding='utf-8') as f:
                            markdown_content = f.read()
                        print(f"Successfully read markdown content, length: {len(markdown_content)}")
                        
                        # Analyze with LLM
                        print("Analyzing with Gemini LLM...")
                        analysis_report = self.llm_analyzer.analyze_markdown(markdown_content)
                        
                        if analysis_report:
                            # Save the analysis report
                            report_path = self.llm_analyzer.save_report(analysis_report, str(workbook_dir))
                            print(f"LLM analysis saved to: {report_path}")
                        else:
                            print("Error: LLM analysis failed")
                    except Exception as e:
                        print(f"Error in LLM analysis: {str(e)}")

# Example usage
if __name__ == "__main__":
    converter = ExcelToLLMConverter(
        input_path=INPUT_DIR,
        output_dir=OUTPUT_DIR,
        api_key="your_api_key_here"  # Pass the API key
    )
    converter.convert_all()
