import pandas as pd
import openpyxl
from pathlib import Path
import json
import re
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple, Any
from datetime import datetime
from combine_markdown import combine_markdown_files
from llm_analyzer import LLMAnalyzer
from prd_generator import PRDGenerator
import os

class EnhancedExcelConverter:
    def __init__(self, input_path: str, output_dir: str, api_key: str, generate_prd: bool = True):
        self.input_path = Path(input_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.llm_analyzer = LLMAnalyzer(api_key)
        self.prd_generator = PRDGenerator(api_key) if generate_prd else None
        self.generate_prd = generate_prd

    def infer_cell_type(self, cell: openpyxl.cell.Cell) -> str:
        """Infer the type of data in a cell."""
        if cell.value is None:
            return "empty"
        
        if cell.data_type == 'n':  # Numeric
            if isinstance(cell.value, datetime):
                return "date"
            if str(cell.number_format).endswith('%'):
                return "percentage"
            if any(currency_symbol in str(cell.number_format) for currency_symbol in ['$', '£', '€', '¥']):
                return "currency"
            return "numeric"
        elif cell.data_type == 'f':
            return "formula"
        elif cell.data_type == 'b':
            return "boolean"
        elif cell.data_type == 's':
            return "text"
        
        return str(cell.data_type)

    def analyze_business_logic_patterns(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Enhanced analysis to identify business logic patterns for PRD generation."""
        patterns = {
            "input_sections": [],
            "calculation_engines": [],
            "output_dashboards": [],
            "validation_rules": [],
            "scenario_controllers": [],
            "data_flow_maps": []
        }
        
        # Identify input sections (cells with data validation or specific styling)
        for row in worksheet.iter_rows():
            for cell in row:
                # Check if cell has data_validation attribute and it's not None
                if hasattr(cell, 'data_validation') and cell.data_validation and hasattr(cell.data_validation, 'type') and cell.data_validation.type:
                    patterns["input_sections"].append({
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "validation_type": cell.data_validation.type,
                        "validation_formula": str(cell.data_validation.formula1) if hasattr(cell.data_validation, 'formula1') and cell.data_validation.formula1 else None,
                        "value": cell.value,
                        "comment": cell.comment.text if cell.comment else None
                    })
        
        # Identify calculation engines (complex formula patterns)
        calculation_areas = []
        for row in worksheet.iter_rows():
            formula_density = sum(1 for cell in row if cell.data_type == 'f')
            if formula_density > 3:  # High formula density indicates calculation area
                calc_area = {
                    "row": row[0].row,
                    "formula_count": formula_density,
                    "formulas": []
                }
                for cell in row:
                    if cell.data_type == 'f':
                        calc_area["formulas"].append({
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "formula": str(cell.value),
                            "category": self.categorize_formula(str(cell.value))
                        })
                calculation_areas.append(calc_area)
        
        patterns["calculation_engines"] = calculation_areas
        
        # Identify output dashboards (formatted sections with charts/summaries)
        dashboard_indicators = ["summary", "dashboard", "report", "total", "analysis"]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = cell.value.lower()
                    if any(indicator in cell_text for indicator in dashboard_indicators):
                        if cell.font and (cell.font.bold or cell.font.size > 12):
                            patterns["output_dashboards"].append({
                                "cell": f"{get_column_letter(cell.column)}{cell.row}",
                                "title": cell.value,
                                "area_start": f"{get_column_letter(cell.column)}{cell.row}"
                            })
        
        return patterns

    def extract_data_dependencies(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract data flow and dependencies for software architecture design."""
        dependencies = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    # Extract cell references
                    cell_refs = re.findall(r'[A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?', formula)
                    sheet_refs = re.findall(r"'?([^'!]+)'?![A-Za-z]+[0-9]+", formula)
                    
                    dependency = {
                        "target_cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "formula": formula,
                        "local_dependencies": cell_refs,
                        "sheet_dependencies": sheet_refs,
                        "dependency_type": self.classify_dependency_type(formula),
                        "complexity_score": self.calculate_formula_complexity(formula)
                    }
                    dependencies.append(dependency)
        
        return dependencies

    def classify_dependency_type(self, formula: str) -> str:
        """Classify the type of dependency for architecture planning."""
        formula_upper = formula.upper()
        
        if '!' in formula:
            return "cross_sheet_reference"
        elif any(func in formula_upper for func in ["VLOOKUP", "INDEX", "MATCH", "XLOOKUP"]):
            return "data_lookup"
        elif any(func in formula_upper for func in ["SUM", "AVERAGE", "COUNT", "MAX", "MIN"]):
            return "aggregation"
        elif "IF(" in formula_upper:
            return "conditional_logic"
        elif any(func in formula_upper for func in ["NPV", "IRR", "PMT", "PV", "FV"]):
            return "financial_calculation"
        else:
            return "simple_calculation"

    def calculate_formula_complexity(self, formula: str) -> int:
        """Calculate complexity score for implementation planning."""
        score = 0
        score += formula.count('(')  # Function calls
        score += formula.count('!')  # Sheet references
        score += formula.count('IF') * 2  # Conditional logic is more complex
        score += len(re.findall(r'[A-Za-z]+[0-9]+', formula))  # Cell references
        return score

    def identify_tables(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Enhanced table identification with business context."""
        tables = []
        current_table = None
        
        for row_idx, row in enumerate(worksheet.iter_rows(), 1):
            # Look for potential header rows
            header_candidates = [cell for cell in row if cell.font and cell.font.bold]
            
            if header_candidates and len(header_candidates) > 1:
                # Found potential new table header
                if current_table:
                    current_table["row_count"] = row_idx - current_table["start_row"]
                    tables.append(current_table)
                
                # Start new table
                start_col = min(cell.column for cell in header_candidates)
                end_col = max(cell.column for cell in header_candidates)
                headers = []
                types = []
                business_context = self.infer_table_business_context(header_candidates)
                
                for cell in row[start_col-1:end_col]:
                    headers.append(str(cell.value) if cell.value else f"Column_{get_column_letter(cell.column)}")
                    types.append(self.infer_cell_type(cell))
                
                current_table = {
                    "name": headers[0] if headers else f"Table_{len(tables)+1}",
                    "range": f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}",
                    "headers": headers,
                    "types": types,
                    "start_row": row_idx,
                    "start_col": start_col,
                    "end_col": end_col,
                    "business_context": business_context,
                    "is_input_table": self.is_input_table(headers),
                    "is_calculation_table": self.is_calculation_table(headers),
                    "is_output_table": self.is_output_table(headers)
                }
        
        # Close last table if exists
        if current_table:
            current_table["row_count"] = worksheet.max_row - current_table["start_row"] + 1
            tables.append(current_table)
        
        return tables

    def infer_table_business_context(self, header_cells: List) -> str:
        """Infer business context from table headers."""
        header_texts = [str(cell.value).lower() for cell in header_cells if cell.value]
        
        financial_keywords = ["revenue", "cost", "profit", "cash", "balance", "income", "expense"]
        operational_keywords = ["volume", "units", "quantity", "capacity", "production"]
        input_keywords = ["assumption", "input", "parameter", "rate", "factor"]
        output_keywords = ["result", "output", "summary", "total", "forecast"]
        
        if any(keyword in ' '.join(header_texts) for keyword in financial_keywords):
            return "financial"
        elif any(keyword in ' '.join(header_texts) for keyword in operational_keywords):
            return "operational"
        elif any(keyword in ' '.join(header_texts) for keyword in input_keywords):
            return "input_parameters"
        elif any(keyword in ' '.join(header_texts) for keyword in output_keywords):
            return "output_results"
        else:
            return "general"

    def is_input_table(self, headers: List[str]) -> bool:
        """Determine if this is an input table."""
        input_indicators = ["input", "assumption", "parameter", "rate", "factor", "variable"]
        header_text = ' '.join(str(h).lower() for h in headers)
        return any(indicator in header_text for indicator in input_indicators)

    def is_calculation_table(self, headers: List[str]) -> bool:
        """Determine if this is a calculation table."""
        calc_indicators = ["calculation", "calc", "formula", "computed", "derived"]
        header_text = ' '.join(str(h).lower() for h in headers)
        return any(indicator in header_text for indicator in calc_indicators)

    def is_output_table(self, headers: List[str]) -> bool:
        """Determine if this is an output table."""
        output_indicators = ["output", "result", "summary", "total", "report", "dashboard"]
        header_text = ' '.join(str(h).lower() for h in headers)
        return any(indicator in header_text for indicator in output_indicators)

    def categorize_formula(self, formula: str) -> str:
        """Enhanced formula categorization."""
        formula = formula.upper()
        if '!' in formula:
            return "external_reference"
        elif any(agg in formula for agg in ["SUM(", "AVERAGE(", "COUNT(", "MAX(", "MIN("]):
            return "aggregation"
        elif "IF(" in formula:
            return "conditional_logic"
        elif any(fin in formula for fin in ["NPV(", "IRR(", "PMT(", "PV(", "FV("]):
            return "financial_function"
        elif any(lookup in formula for lookup in ["VLOOKUP(", "INDEX(", "MATCH(", "XLOOKUP("]):
            return "data_lookup"
        return "other"

    def extract_named_ranges(self, workbook: openpyxl.Workbook) -> List[Dict[str, str]]:
        """Extract named ranges with enhanced metadata."""
        named_ranges = []
        try:
            if hasattr(workbook, 'defined_names'):
                for name, defn in workbook.defined_names.items():
                    try:
                        destinations = defn.destinations
                        for worksheet, coordinate in destinations:
                            named_ranges.append({
                                "name": name,
                                "range": f"{worksheet}!{coordinate}",
                                "business_purpose": self.infer_named_range_purpose(name)
                            })
                    except AttributeError:
                        named_ranges.append({
                            "name": name,
                            "range": str(defn.value),
                            "business_purpose": self.infer_named_range_purpose(name)
                        })
        except Exception as e:
            print(f"Warning: Could not extract named ranges: {str(e)}")
        
        return named_ranges

    def infer_named_range_purpose(self, name: str) -> str:
        """Infer business purpose from named range name."""
        name_lower = name.lower()
        
        if any(word in name_lower for word in ["input", "param", "assumption"]):
            return "input_parameter"
        elif any(word in name_lower for word in ["rate", "factor", "multiplier"]):
            return "calculation_factor"
        elif any(word in name_lower for word in ["output", "result", "total"]):
            return "output_value"
        elif any(word in name_lower for word in ["scenario", "case", "option"]):
            return "scenario_control"
        else:
            return "general"

    def process_worksheet(self, worksheet: Worksheet, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Enhanced worksheet processing with PRD-focused analysis."""
        sheet_data = {
            "name": worksheet.title,
            "dimensions": f"{worksheet.dimensions}",
            "tables": self.identify_tables(worksheet),
            "named_ranges": self.extract_named_ranges(workbook),
            "business_logic_patterns": self.analyze_business_logic_patterns(worksheet),
            "data_dependencies": self.extract_data_dependencies(worksheet),
            "cells": {},
            "formulas": {
                "external_references": [],
                "aggregations": [],
                "conditional_logic": [],
                "financial_functions": [],
                "data_lookups": [],
                "other": []
            },
            "software_requirements": {
                "ui_components": self.identify_ui_components(worksheet),
                "business_rules": self.extract_business_rules(worksheet),
                "data_validation_rules": self.extract_validation_rules(worksheet),
                "calculation_sequences": self.identify_calculation_sequences(worksheet)
            }
        }

        # Process each cell with enhanced metadata
        for row in worksheet.iter_rows():
            for cell in row:
                cell_addr = f"{get_column_letter(cell.column)}{cell.row}"
                
                # Basic cell data
                cell_data = {
                    "value": cell.value,
                    "type": self.infer_cell_type(cell),
                    "has_formula": cell.data_type == 'f',
                    "is_styled": bool(cell.font and (cell.font.bold or cell.font.italic)),
                    "has_validation": bool(hasattr(cell, 'data_validation') and cell.data_validation and hasattr(cell.data_validation, 'type') and cell.data_validation.type),
                    "has_comment": bool(cell.comment),
                    "business_context": self.infer_cell_business_context(cell)
                }
                
                # Add to appropriate formula category
                if cell.data_type == 'f':
                    formula_metadata = self.extract_formula_metadata(cell)
                    category = formula_metadata["category"]
                    if category in sheet_data["formulas"]:
                        sheet_data["formulas"][category].append(formula_metadata)
                
                sheet_data["cells"][cell_addr] = cell_data

        return sheet_data

    def infer_cell_business_context(self, cell) -> str:
        """Infer business context of individual cells."""
        if cell.comment:
            return "documented"
        elif hasattr(cell, 'data_validation') and cell.data_validation and hasattr(cell.data_validation, 'type') and cell.data_validation.type:
            return "user_input"
        elif cell.data_type == 'f':
            return "calculated"
        elif cell.font and cell.font.bold:
            return "header_or_label"
        else:
            return "data"

    def identify_ui_components(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Identify UI components needed for software implementation."""
        components = []
        
        # Input components (cells with data validation)
        for row in worksheet.iter_rows():
            for cell in row:
                if hasattr(cell, 'data_validation') and cell.data_validation and hasattr(cell.data_validation, 'type') and cell.data_validation.type:
                    component = {
                        "type": "input_field",
                        "location": f"{get_column_letter(cell.column)}{cell.row}",
                        "input_type": cell.data_validation.type,
                        "validation": str(cell.data_validation.formula1) if hasattr(cell.data_validation, 'formula1') and cell.data_validation.formula1 else None,
                        "current_value": cell.value,
                        "label": self.find_cell_label(worksheet, cell)
                    }
                    components.append(component)
        
        # Output components (formatted display areas)
        for table in self.identify_tables(worksheet):
            if table["is_output_table"]:
                component = {
                    "type": "data_table",
                    "location": table["range"],
                    "headers": table["headers"],
                    "business_context": table["business_context"]
                }
                components.append(component)
        
        return components

    def find_cell_label(self, worksheet: Worksheet, target_cell) -> str:
        """Find the label for an input cell."""
        # Check cells to the left and above
        row, col = target_cell.row, target_cell.column
        
        # Check left
        if col > 1:
            left_cell = worksheet.cell(row, col - 1)
            if left_cell.value and isinstance(left_cell.value, str):
                return left_cell.value
        
        # Check above
        if row > 1:
            above_cell = worksheet.cell(row - 1, col)
            if above_cell.value and isinstance(above_cell.value, str):
                return above_cell.value
        
        return f"Cell_{get_column_letter(col)}{row}"

    def extract_business_rules(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract business rules from formulas and patterns."""
        rules = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    if "IF(" in formula.upper():
                        rule = {
                            "type": "conditional_rule",
                            "location": f"{get_column_letter(cell.column)}{cell.row}",
                            "formula": formula,
                            "description": self.describe_conditional_logic(formula)
                        }
                        rules.append(rule)
        
        return rules

    def describe_conditional_logic(self, formula: str) -> str:
        """Convert IF formula to business rule description."""
        # Simplified description generation
        if "IF(" in formula.upper():
            return f"Conditional calculation based on: {formula}"
        return formula

    def extract_validation_rules(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Extract data validation rules for software implementation."""
        validation_rules = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if hasattr(cell, 'data_validation') and cell.data_validation and hasattr(cell.data_validation, 'type') and cell.data_validation.type:
                    rule = {
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "validation_type": cell.data_validation.type,
                        "formula": str(cell.data_validation.formula1) if hasattr(cell.data_validation, 'formula1') and cell.data_validation.formula1 else None,
                        "error_message": cell.data_validation.error if hasattr(cell.data_validation, 'error') and cell.data_validation.error else None,
                        "input_message": cell.data_validation.prompt if hasattr(cell.data_validation, 'prompt') and cell.data_validation.prompt else None
                    }
                    validation_rules.append(rule)
        
        return validation_rules

    def identify_calculation_sequences(self, worksheet: Worksheet) -> List[Dict[str, Any]]:
        """Identify sequences of calculations for implementation planning."""
        sequences = []
        dependencies = self.extract_data_dependencies(worksheet)
        
        # Group related calculations
        calc_groups = {}
        for dep in dependencies:
            if dep["dependency_type"] not in calc_groups:
                calc_groups[dep["dependency_type"]] = []
            calc_groups[dep["dependency_type"]].append(dep)
        
        for calc_type, calcs in calc_groups.items():
            sequence = {
                "type": calc_type,
                "calculations": calcs,
                "complexity": sum(c["complexity_score"] for c in calcs),
                "implementation_priority": self.assess_implementation_priority(calc_type, calcs)
            }
            sequences.append(sequence)
        
        return sequences

    def assess_implementation_priority(self, calc_type: str, calculations: List) -> str:
        """Assess implementation priority for calculation sequences."""
        priority_map = {
            "simple_calculation": "low",
            "aggregation": "medium",
            "conditional_logic": "medium",
            "data_lookup": "high",
            "financial_calculation": "high",
            "cross_sheet_reference": "high"
        }
        return priority_map.get(calc_type, "medium")

    def extract_formula_metadata(self, cell: openpyxl.cell.Cell) -> Dict[str, Any]:
        """Enhanced formula metadata extraction."""
        metadata = {
            "address": f"{get_column_letter(cell.column)}{cell.row}",
            "value": cell.value,
            "data_type": self.infer_cell_type(cell),
            "formula": None,
            "category": None,
            "dependencies": [],
            "complexity_score": 0,
            "implementation_notes": ""
        }

        if cell.data_type == 'f':
            formula = str(cell.value)
            if formula.startswith('='):
                metadata["formula"] = formula
                metadata["category"] = self.categorize_formula(formula)
                metadata["dependencies"] = re.findall(r'[A-Za-z]+[0-9]+(?::[A-Za-z]+[0-9]+)?', formula)
                metadata["complexity_score"] = self.calculate_formula_complexity(formula)
                metadata["implementation_notes"] = self.generate_implementation_notes(formula)

        return metadata

    def generate_implementation_notes(self, formula: str) -> str:
        """Generate implementation notes for formulas."""
        formula_upper = formula.upper()
        notes = []
        
        if "VLOOKUP(" in formula_upper:
            notes.append("Requires database lookup functionality")
        if "IF(" in formula_upper:
            notes.append("Implement conditional logic with proper error handling")
        if "SUM(" in formula_upper:
            notes.append("Use efficient aggregation queries")
        if "!" in formula:
            notes.append("Requires cross-table/cross-module data access")
        
        return "; ".join(notes) if notes else "Standard calculation implementation"

    def sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for cross-platform compatibility."""
        return re.sub(r'[<>:"/\\|?*]', '_', filename)

    def convert_to_markdown(self, sheet_data: Dict[str, Any], output_file: Path) -> None:
        """Enhanced markdown conversion with PRD-focused structure."""
        with output_file.open('w', encoding='utf-8') as f:
            # Write header
            f.write(f"# Sheet: {sheet_data['name']}\n\n")
            f.write(f"Dimensions: {sheet_data['dimensions']}\n\n")

            # Software Requirements Section
            f.write("## Software Implementation Requirements\n\n")
            
            # UI Components
            if sheet_data["software_requirements"]["ui_components"]:
                f.write("### UI Components Required\n\n")
                for component in sheet_data["software_requirements"]["ui_components"]:
                    f.write(f"- **{component['type']}** at {component['location']}\n")
                    if component.get('label'):
                        f.write(f"  - Label: {component['label']}\n")
                    if component.get('validation'):
                        f.write(f"  - Validation: {component['validation']}\n")
                    f.write("\n")

            # Business Rules
            if sheet_data["software_requirements"]["business_rules"]:
                f.write("### Business Rules\n\n")
                for rule in sheet_data["software_requirements"]["business_rules"]:
                    f.write(f"- **{rule['type']}** at {rule['location']}\n")
                    f.write(f"  - Description: {rule['description']}\n\n")

            # Calculation Sequences
            if sheet_data["software_requirements"]["calculation_sequences"]:
                f.write("### Calculation Implementation Sequences\n\n")
                for seq in sheet_data["software_requirements"]["calculation_sequences"]:
                    f.write(f"#### {seq['type'].replace('_', ' ').title()}\n")
                    f.write(f"- Complexity Score: {seq['complexity']}\n")
                    f.write(f"- Implementation Priority: {seq['implementation_priority']}\n")
                    f.write(f"- Number of Calculations: {len(seq['calculations'])}\n\n")

            # Original sections (enhanced)
            if sheet_data["tables"]:
                f.write("## Data Tables\n\n")
                for table in sheet_data["tables"]:
                    f.write(f"### {table['name']}\n")
                    f.write(f"- Range: {table['range']}\n")
                    f.write(f"- Headers: {', '.join(table['headers'])}\n")
                    f.write(f"- Business Context: {table['business_context']}\n")
                    f.write(f"- Input Table: {table['is_input_table']}\n")
                    f.write(f"- Calculation Table: {table['is_calculation_table']}\n")
                    f.write(f"- Output Table: {table['is_output_table']}\n")
                    f.write("\n")

            # Data Dependencies
            if sheet_data["data_dependencies"]:
                f.write("## Data Flow and Dependencies\n\n")
                dependency_types = {}
                for dep in sheet_data["data_dependencies"]:
                    dep_type = dep["dependency_type"]
                    if dep_type not in dependency_types:
                        dependency_types[dep_type] = []
                    dependency_types[dep_type].append(dep)
                
                for dep_type, deps in dependency_types.items():
                    f.write(f"### {dep_type.replace('_', ' ').title()}\n")
                    for dep in deps[:5]:  # Limit to first 5 for readability
                        f.write(f"- {dep['target_cell']}: {dep['formula'][:50]}{'...' if len(dep['formula']) > 50 else ''}\n")
                        f.write(f"  - Complexity: {dep['complexity_score']}\n")
                    if len(deps) > 5:
                        f.write(f"  - ... and {len(deps) - 5} more\n")
                    f.write("\n")

            # Continue with existing sections...
            if sheet_data["named_ranges"]:
                f.write("## Named Ranges\n\n")
                for named_range in sheet_data["named_ranges"]:
                    f.write(f"- {named_range['name']}: {named_range['range']}")
                    if named_range.get('business_purpose'):
                        f.write(f" (Purpose: {named_range['business_purpose']})")
                    f.write("\n")

            # Enhanced formulas section
            if any(sheet_data["formulas"].values()):
                f.write("\n## Formulas by Category\n\n")
                for category, formulas in sheet_data["formulas"].items():
                    if formulas:
                        f.write(f"### {category.replace('_', ' ').title()}\n")
                        for formula in formulas[:10]:  # Limit for readability
                            f.write(f"- {formula['address']}: `{formula['formula']}`\n")
                            if formula.get('implementation_notes'):
                                f.write(f"  - Implementation: {formula['implementation_notes']}\n")
                            if formula['dependencies']:
                                f.write(f"  - Dependencies: {', '.join(formula['dependencies'])}\n")
                        if len(formulas) > 10:
                            f.write(f"  - ... and {len(formulas) - 10} more formulas\n")
                        f.write("\n")

    def generate_workbook_summary(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Enhanced workbook summary with PRD-focused metadata."""
        summary = {
            "sheet_count": len(workbook.worksheets),
            "sheets": [],
            "formula_patterns": {},
            "most_formulas": {"sheet": None, "count": 0},
            "business_complexity": {
                "total_input_fields": 0,
                "total_calculations": 0,
                "total_outputs": 0,
                "cross_sheet_references": 0
            },
            "implementation_estimates": {
                "ui_components": 0,
                "business_rules": 0,
                "integrations_needed": 0,
                "complexity_score": 0
            }
        }

        for worksheet in workbook.worksheets:
            sheet_data = self.process_worksheet(worksheet, workbook)
            
            sheet_summary = {
                "name": worksheet.title,
                "formula_count": 0,
                "table_count": len(sheet_data["tables"]),
                "ui_components": len(sheet_data["software_requirements"]["ui_components"]),
                "business_rules": len(sheet_data["software_requirements"]["business_rules"]),
                "calculation_sequences": len(sheet_data["software_requirements"]["calculation_sequences"])
            }

            # Count formulas and patterns
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        sheet_summary["formula_count"] += 1
                        formula = str(cell.value).upper()
                        
                        # Count cross-sheet references
                        if '!' in formula:
                            summary["business_complexity"]["cross_sheet_references"] += 1
                        
                        # Pattern analysis
                        for pattern in ["SUM(", "IF(", "VLOOKUP(", "INDEX(", "MATCH(", "NPV(", "IRR("]:
                            if pattern in formula:
                                summary["formula_patterns"][pattern] = summary["formula_patterns"].get(pattern, 0) + 1

            # Update summary counts
            summary["implementation_estimates"]["ui_components"] += sheet_summary["ui_components"]
            summary["implementation_estimates"]["business_rules"] += sheet_summary["business_rules"]
            summary["implementation_estimates"]["complexity_score"] += sheet_summary["formula_count"]

            summary["sheets"].append(sheet_summary)
            
            if sheet_summary["formula_count"] > summary["most_formulas"]["count"]:
                summary["most_formulas"] = {
                    "sheet": worksheet.title,
                    "count": sheet_summary["formula_count"]
                }

        # Calculate overall complexity
        total_formulas = sum(s["formula_count"] for s in summary["sheets"])
        if total_formulas > 500:
            summary["complexity_rating"] = "Very High"
        elif total_formulas > 200:
            summary["complexity_rating"] = "High"
        elif total_formulas > 50:
            summary["complexity_rating"] = "Medium"
        else:
            summary["complexity_rating"] = "Low"

        # Sort formula patterns by frequency
        summary["formula_patterns"] = dict(
            sorted(summary["formula_patterns"].items(), key=lambda x: x[1], reverse=True)
        )

        return summary

    def process_workbook(self, excel_file: Path) -> None:
        """Enhanced workbook processing with PRD generation."""
        try:
            print(f"Processing {excel_file}...")
            workbook = openpyxl.load_workbook(excel_file, data_only=False)

            # Generate enhanced workbook summary
            workbook_summary = self.generate_workbook_summary(workbook)

            # Create output directory for this workbook
            workbook_dir = self.output_dir / excel_file.stem
            workbook_dir.mkdir(exist_ok=True)

            # Save enhanced workbook summary
            self.save_enhanced_workbook_summary(workbook_summary, workbook_dir)

            # Process each worksheet
            for worksheet in workbook.worksheets:
                print(f"Processing worksheet: {worksheet.title}")
                sheet_data = self.process_worksheet(worksheet, workbook)

                # Sanitize the worksheet title for filename
                safe_title = self.sanitize_filename(worksheet.title)
                if not safe_title:
                    safe_title = "Sheet"

                # Save as enhanced markdown
                md_file = workbook_dir / f"{safe_title}.md"
                self.convert_to_markdown(sheet_data, md_file)
                print(f"Created enhanced markdown file: {md_file}")

                # Save enhanced JSON metadata
                json_file = workbook_dir / f"{safe_title}.json"
                with json_file.open('w', encoding='utf-8') as f:
                    json.dump(sheet_data, f, indent=2, default=str)
                print(f"Created enhanced JSON file: {json_file}")

        except Exception as e:
            print(f"Error processing {excel_file}: {str(e)}")

    def save_enhanced_workbook_summary(self, workbook_summary: Dict[str, Any], workbook_dir: Path) -> None:
        """Save enhanced workbook summary with implementation insights."""
        with (workbook_dir / "enhanced_workbook_summary.md").open('w', encoding='utf-8') as f:
            f.write("# Enhanced Workbook Summary\n\n")
            f.write(f"Total Sheets: {workbook_summary['sheet_count']}\n")
            f.write(f"Overall Complexity: {workbook_summary['complexity_rating']}\n\n")
            
            f.write("## Implementation Estimates\n")
            f.write(f"- UI Components Required: {workbook_summary['implementation_estimates']['ui_components']}\n")
            f.write(f"- Business Rules to Implement: {workbook_summary['implementation_estimates']['business_rules']}\n")
            f.write(f"- Total Complexity Score: {workbook_summary['implementation_estimates']['complexity_score']}\n")
            f.write(f"- Cross-Sheet References: {workbook_summary['business_complexity']['cross_sheet_references']}\n\n")
            
            f.write("## Sheet Details\n")
            for sheet in workbook_summary["sheets"]:
                f.write(f"\n### {sheet['name']}\n")
                f.write(f"- Formula Count: {sheet['formula_count']}\n")
                f.write(f"- Table Count: {sheet['table_count']}\n")
                f.write(f"- UI Components: {sheet['ui_components']}\n")
                f.write(f"- Business Rules: {sheet['business_rules']}\n")
                f.write(f"- Calculation Sequences: {sheet['calculation_sequences']}\n")
            
            f.write("\n## Formula Analysis\n")
            f.write(f"Sheet with Most Formulas: {workbook_summary['most_formulas']['sheet']} ")
            f.write(f"({workbook_summary['most_formulas']['count']} formulas)\n\n")
            
            if workbook_summary["formula_patterns"]:
                f.write("Common Formula Patterns:\n")
                for pattern, count in workbook_summary["formula_patterns"].items():
                    f.write(f"- {pattern}: {count} occurrences\n")

    def convert_all(self):
        """Enhanced conversion with PRD generation."""
        if self.input_path.is_file():
            self.process_workbook(self.input_path)
        else:
            for excel_file in self.input_path.glob("*.xlsx"):
                self.process_workbook(excel_file)
        
        # After processing all Excel files, generate combined analysis and PRD
        for workbook_dir in self.output_dir.iterdir():
            if workbook_dir.is_dir():
                print(f"\nCombining markdown files for {workbook_dir.name}...")
                combined_file = combine_markdown_files(str(workbook_dir), "combined_enhanced_workbook.md")
                
                if combined_file and os.path.exists(combined_file):
                    print(f"Successfully created combined markdown file: {combined_file}")
                    
                    try:
                        with open(combined_file, 'r', encoding='utf-8') as f:
                            markdown_content = f.read()
                        print(f"Successfully read markdown content, length: {len(markdown_content)}")
                        
                        # Generate user guide with LLM analyzer
                        print("Analyzing with Gemini LLM for user guide...")
                        analysis_report = self.llm_analyzer.analyze_markdown(markdown_content)
                        
                        if analysis_report:
                            report_path = self.llm_analyzer.save_report(analysis_report, str(workbook_dir))
                            print(f"User guide analysis saved to: {report_path}")
                        
                        # Generate PRD if enabled
                        if self.generate_prd and self.prd_generator:
                            print("Generating Product Requirements Document...")
                            
                            # Extract metadata for enhanced PRD generation
                            summary_path = workbook_dir / "enhanced_workbook_summary.md"
                            metadata = self.prd_generator.extract_spreadsheet_metadata(str(summary_path))
                            
                            prd_content = self.prd_generator.generate_prd(markdown_content, metadata)
                            
                            if prd_content:
                                prd_path = self.prd_generator.save_prd(prd_content, str(workbook_dir))
                                print(f"PRD document saved to: {prd_path}")
                            else:
                                print("Error: PRD generation failed")
                        
                    except Exception as e:
                        print(f"Error in analysis: {str(e)}")

# Example usage
if __name__ == "__main__":
    converter = EnhancedExcelConverter(
        input_path="your_input_path",
        output_dir="your_output_path",
        api_key="your_api_key_here",
        generate_prd=True
    )
    converter.convert_all() 